#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl


def log(message: str) -> None:
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {message}", flush=True)


def normalize_sku(value: object) -> str:
    if value is None:
        return ""
    return "".join(str(value).split())


def parse_quantity(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def list_official_files(official_dir: Path) -> List[Path]:
    if not official_dir.exists():
        return []

    files: List[Path] = []
    for path in official_dir.iterdir():
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() != ".xlsx":
            continue
        files.append(path)
    return sorted(files)


def find_newest_csv(site_dir: Path) -> Path:
    if not site_dir.exists():
        raise FileNotFoundError(f"Site directory does not exist: {site_dir}")

    csv_files = [p for p in site_dir.iterdir() if p.is_file() and p.suffix.lower() == ".csv"]
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in site directory: {site_dir}")

    return max(csv_files, key=lambda p: p.stat().st_mtime_ns)


def parse_official_workbook(path: Path) -> Tuple[Dict[str, int], Dict[str, int]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    qty_by_sku: Dict[str, float] = {}
    count_by_sku: Dict[str, int] = {}
    total_rows = 0

    try:
        for row in ws.iter_rows(values_only=True):
            total_rows += 1
            sku_raw = row[0] if len(row) > 0 else None
            qty_raw = row[2] if len(row) > 2 else None

            sku = normalize_sku(sku_raw)
            if not sku:
                continue

            qty = parse_quantity(qty_raw)
            qty_by_sku[sku] = qty_by_sku.get(sku, 0.0) + qty
            count_by_sku[sku] = count_by_sku.get(sku, 0) + 1
    finally:
        wb.close()

    availability = {sku: 1 if qty > 0 else 0 for sku, qty in qty_by_sku.items()}
    duplicate_keys = sum(1 for count in count_by_sku.values() if count > 1)

    stats = {
        "official_rows": total_rows,
        "unique_skus": len(availability),
        "duplicate_sku_keys": duplicate_keys,
    }
    return availability, stats


def update_site_csv(csv_path: Path, availability: Dict[str, int]) -> Dict[str, int]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        fieldnames = reader.fieldnames or []

        required = {"Артикул", "Наличие"}
        missing = sorted(required - set(fieldnames))
        if missing:
            raise ValueError(
                f"CSV {csv_path} is missing required columns: {', '.join(missing)}"
            )

        rows = list(reader)

    matched_rows = 0
    set_to_1 = 0
    set_to_0 = 0

    for row in rows:
        sku = normalize_sku(row.get("Артикул", ""))
        if sku in availability:
            value = availability[sku]
            matched_rows += 1
        else:
            value = 0

        row["Наличие"] = str(value)
        if value == 1:
            set_to_1 += 1
        else:
            set_to_0 += 1

    fd, tmp_path_str = tempfile.mkstemp(
        prefix=f"{csv_path.name}.", suffix=".tmp", dir=str(csv_path.parent)
    )
    os.close(fd)
    tmp_path = Path(tmp_path_str)

    try:
        with tmp_path.open("w", encoding="utf-8-sig", newline="") as target:
            writer = csv.DictWriter(target, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        os.replace(tmp_path, csv_path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()

    return {
        "site_rows": len(rows),
        "matched_rows": matched_rows,
        "set_to_1": set_to_1,
        "set_to_0": set_to_0,
    }


def process_single_official_file(official_xlsx: Path, site_dir: Path) -> Dict[str, object]:
    availability, official_stats = parse_official_workbook(official_xlsx)
    site_csv = find_newest_csv(site_dir)
    csv_stats = update_site_csv(site_csv, availability)

    summary: Dict[str, object] = {
        "official_file": str(official_xlsx),
        "site_csv": str(site_csv),
    }
    summary.update(official_stats)
    summary.update(csv_stats)
    return summary


def fingerprint(path: Path) -> str:
    st = path.stat()
    return f"{st.st_mtime_ns}:{st.st_size}"


def load_state(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}

    processed = data.get("processed", {})
    if not isinstance(processed, dict):
        return {}

    normalized: Dict[str, str] = {}
    for key, value in processed.items():
        if isinstance(key, str) and isinstance(value, str):
            normalized[key] = value
    return normalized


def save_state(path: Path, processed: Dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"processed": processed}

    fd, tmp_path_str = tempfile.mkstemp(
        prefix=f"{path.name}.", suffix=".tmp", dir=str(path.parent)
    )
    os.close(fd)
    tmp_path = Path(tmp_path_str)

    try:
        tmp_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        os.replace(tmp_path, path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def run_once(official_dir: Path, site_dir: Path) -> Dict[str, object]:
    files = list_official_files(official_dir)
    if not files:
        raise FileNotFoundError(f"No official XLSX files found in: {official_dir}")

    newest = max(files, key=lambda p: p.stat().st_mtime_ns)
    summary = process_single_official_file(newest, site_dir)
    log(
        "Processed once | official={official_file} | site={site_csv} | "
        "rows={official_rows} unique={unique_skus} dup={duplicate_sku_keys} "
        "matched={matched_rows} set_to_1={set_to_1} set_to_0={set_to_0}".format(**summary)
    )
    return summary


def run_watch(
    official_dir: Path,
    site_dir: Path,
    poll_interval: float,
    debounce_seconds: float,
    state_path: Path,
) -> None:
    processed = load_state(state_path)
    observed: Dict[str, Dict[str, object]] = {}

    log(
        f"Watching {official_dir} every {poll_interval}s "
        f"(debounce={debounce_seconds}s, state={state_path})"
    )

    while True:
        files = list_official_files(official_dir)
        current_keys = set()
        now = time.monotonic()

        for path in files:
            key = str(path.resolve())
            current_keys.add(key)

            try:
                current_fp = fingerprint(path)
            except OSError as exc:
                log(f"Skipping {path}: cannot stat file ({exc})")
                continue

            prior = observed.get(key)
            if prior is None or prior.get("fingerprint") != current_fp:
                observed[key] = {
                    "fingerprint": current_fp,
                    "first_seen": now,
                }
                continue

            stable_for = now - float(prior["first_seen"])
            if stable_for < debounce_seconds:
                continue

            if processed.get(key) == current_fp:
                continue

            try:
                summary = process_single_official_file(path, site_dir)
                processed[key] = current_fp
                save_state(state_path, processed)
                log(
                    "Processed | official={official_file} | site={site_csv} | "
                    "rows={official_rows} unique={unique_skus} dup={duplicate_sku_keys} "
                    "matched={matched_rows} set_to_1={set_to_1} set_to_0={set_to_0}".format(
                        **summary
                    )
                )
            except Exception as exc:
                log(f"ERROR processing {path}: {exc}")

        stale_keys = [key for key in observed if key not in current_keys]
        for key in stale_keys:
            observed.pop(key, None)

        time.sleep(poll_interval)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Watch official stock XLSX files and sync availability into site CSV."
    )
    parser.add_argument("--official-dir", default="stock_official", help="Official XLSX folder")
    parser.add_argument("--site-dir", default="stock_site", help="Site CSV folder")
    parser.add_argument(
        "--poll-interval",
        type=float,
        default=5.0,
        help="Polling interval in seconds (watch mode)",
    )
    parser.add_argument(
        "--debounce-seconds",
        type=float,
        default=3.0,
        help="File stability debounce in seconds (watch mode)",
    )
    parser.add_argument(
        "--state-file",
        default=".stock_sync_state.json",
        help="Path for persisted watcher state",
    )
    parser.add_argument(
        "--once",
        action="store_true",
        help="Process newest official XLSX once and exit",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    official_dir = Path(args.official_dir).expanduser().resolve()
    site_dir = Path(args.site_dir).expanduser().resolve()
    state_path = Path(args.state_file).expanduser().resolve()

    try:
        if args.once:
            run_once(official_dir, site_dir)
            return 0

        run_watch(
            official_dir=official_dir,
            site_dir=site_dir,
            poll_interval=args.poll_interval,
            debounce_seconds=args.debounce_seconds,
            state_path=state_path,
        )
        return 0
    except KeyboardInterrupt:
        log("Stopped by user")
        return 0
    except Exception as exc:
        log(f"FATAL: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

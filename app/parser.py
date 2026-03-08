from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from itertools import chain
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

import openpyxl


HEADER_KEYWORDS = {
    "sku",
    "артикул",
    "код",
    "code",
    "qty",
    "quantity",
    "количество",
    "остаток",
    "stock",
    "цена",
    "price",
}


@dataclass
class ParsedItem:
    sku: str
    quantity: float
    availability: int
    price: Optional[str]


@dataclass
class ParseResult:
    stats: Dict[str, int]
    items: Dict[str, ParsedItem]


def normalize_sku(value: object) -> str:
    if value is None:
        return ""
    return "".join(str(value).split())


def _clean_numeric_text(value: str) -> str:
    text = value.replace("\u00a0", " ")
    text = "".join(text.split())
    text = text.replace(",", ".")
    return text


def parse_quantity_strict(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    cleaned = _clean_numeric_text(text)
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_quantity(value: object) -> float:
    parsed = parse_quantity_strict(value)
    if parsed is None:
        return 0.0
    return parsed


def normalize_price(value: object) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value))

    text = str(value).strip()
    if not text:
        return None

    cleaned = _clean_numeric_text(text)
    if cleaned in {"", "-", ".", "-."}:
        return None

    if cleaned.count(".") > 1:
        return None

    try:
        number = Decimal(cleaned)
    except InvalidOperation:
        return None

    return str(int(number))


def _row_values(row: Iterable[object]) -> Tuple[object, object, object, object]:
    values = list(row)
    values.extend([None] * (4 - len(values)))
    return values[0], values[1], values[2], values[3]


def detect_header_row(row: Iterable[object]) -> bool:
    col_a, _col_b, col_c, col_d = _row_values(row)
    cells = [col_a, col_c, col_d]

    for cell in cells:
        if cell is None:
            continue
        text = str(cell).strip().lower()
        if not text:
            continue
        if any(keyword in text for keyword in HEADER_KEYWORDS):
            return True

    return False


def parse_official_xlsx(path: Path) -> ParseResult:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    qty_by_sku: Dict[str, float] = {}
    price_by_sku: Dict[str, Optional[str]] = {}
    count_by_sku: Dict[str, int] = {}

    total_rows = 0
    data_rows = 0
    skipped_empty_sku = 0
    invalid_price_rows = 0

    try:
        iterator = ws.iter_rows(values_only=True)
        first_row = next(iterator, None)
        if first_row is None:
            return ParseResult(
                stats={
                    "total_rows": 0,
                    "data_rows": 0,
                    "skipped_empty_sku": 0,
                    "invalid_price_rows": 0,
                    "unique_skus": 0,
                    "duplicate_sku_keys": 0,
                },
                items={},
            )

        if detect_header_row(first_row):
            total_rows += 1
            rows_to_process = iterator
        else:
            rows_to_process = chain((first_row,), iterator)

        for row in rows_to_process:
            total_rows += 1
            col_a, _col_b, col_c, col_d = _row_values(row)
            sku = normalize_sku(col_a)
            if not sku:
                skipped_empty_sku += 1
                continue

            data_rows += 1
            qty_by_sku[sku] = qty_by_sku.get(sku, 0.0) + parse_quantity(col_c)
            count_by_sku[sku] = count_by_sku.get(sku, 0) + 1

            parsed_price = normalize_price(col_d)
            if col_d is not None and str(col_d).strip() and parsed_price is None:
                invalid_price_rows += 1

            if parsed_price is not None:
                price_by_sku[sku] = parsed_price
            elif sku not in price_by_sku:
                price_by_sku[sku] = None
    finally:
        wb.close()

    duplicate_sku_keys = sum(1 for v in count_by_sku.values() if v > 1)
    items: Dict[str, ParsedItem] = {}
    for sku, quantity in qty_by_sku.items():
        items[sku] = ParsedItem(
            sku=sku,
            quantity=quantity,
            availability=1 if quantity > 0 else 0,
            price=price_by_sku.get(sku),
        )

    stats = {
        "total_rows": total_rows,
        "data_rows": data_rows,
        "skipped_empty_sku": skipped_empty_sku,
        "invalid_price_rows": invalid_price_rows,
        "unique_skus": len(items),
        "duplicate_sku_keys": duplicate_sku_keys,
    }
    return ParseResult(stats=stats, items=items)
